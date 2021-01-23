import openpyxl
import os


# 写excel
def write_excel():
    f = openpyxl.Workbook()  # 创建工作簿

    # sheet1 = f.get_sheet_by_name(f.sheetnames[0])
    sheet1 = f.worksheets[0]
    # print()

    # sheet1 = f.add_sheet(u'sheet1',cell_overwrite_ok=True) #创建sheet
    row0 = [u'L1', u'L2', u'L3', u'L4', u'问题', u'答案']

    # 生成第一行
    for i in range(len(row0)):
        sheet1.cell(column=i + 1, row=1).value = row0[i]
    # for col in sheet1.columns:
    #     print(col)
    print(sheet1.cell(1, 1).value)
    print("===========")
    print(sheet1.max_column)
    print(sheet1.max_row)
    print("===========")
    for row_index in range(1, sheet1.max_row + 1):
        for col_index in range(1, sheet1.max_column+1):
            print(sheet1.cell(row_index , col_index).value)
    # 生成后续

    # for jkey in range(len(newTables)):
    #     jk = 1
    #     for cT in range(arrayNum):
    #         jk = jkey + 1
    #         if cT == 0:
    #             sheet1.cell(row=jk, column=cT + 1).value = '1'
    #         else:
    #             sheet1.cell(row=jk, column=cT + 1).value = '2'

    f.save("chatPy2.xlsx")  # 保存文件
    print(f)
    pass


if __name__ == '__main__':
    print(os.path)
    # 写入Excel
    write_excel()
    print('写入成功')
